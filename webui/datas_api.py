# encoding: utf-8
"""数据目录浏览：扫描已保存的 Excel / 导出、读取 Excel 内容、在访达中打开。"""

import json
import os
import subprocess
import sys
import time

import openpyxl

DATAS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../datas"))
EXCEL_DIR = os.path.join(DATAS_DIR, "excel_datas")
MEDIA_DIR = os.path.join(DATAS_DIR, "media_datas")
EXPORT_DIR = os.path.join(DATAS_DIR, "exports")
HOTSPOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../data/hotspot"))

EXCEL_HEADERS = [
    "笔记id", "笔记url", "笔记类型", "用户id", "用户主页url", "昵称", "头像url",
    "标题", "描述", "点赞数量", "收藏数量", "评论数量", "分享数量",
    "视频封面url", "视频地址url", "图片地址url列表", "标签", "上传时间", "ip归属地",
]


def _mtime(path) -> float:
    try:
        return os.path.getmtime(path)
    except OSError:
        return 0


def scan_datas() -> dict:
    """扫描 datas 目录，返回结构化的数据清单。"""
    excels = []
    if os.path.isdir(EXCEL_DIR):
        for name in sorted(os.listdir(EXCEL_DIR), key=lambda n: _mtime(os.path.join(EXCEL_DIR, n)), reverse=True):
            path = os.path.join(EXCEL_DIR, name)
            if name.endswith(".xlsx") and os.path.isfile(path) and not name.startswith("."):
                excels.append({
                    "name": name,
                    "path": path,
                    "mtime": _mtime(path),
                    "time": time.strftime("%Y-%m-%d %H:%M", time.localtime(_mtime(path))),
                    "size": os.path.getsize(path),
                })

    media_groups = []
    if os.path.isdir(MEDIA_DIR):
        for name in sorted(os.listdir(MEDIA_DIR), key=lambda n: _mtime(os.path.join(MEDIA_DIR, n)), reverse=True):
            path = os.path.join(MEDIA_DIR, name)
            if os.path.isdir(path):
                try:
                    note_count = len([d for d in os.listdir(path) if os.path.isdir(os.path.join(path, d))])
                except OSError:
                    note_count = 0
                media_groups.append({
                    "name": name,
                    "path": path,
                    "mtime": _mtime(path),
                    "time": time.strftime("%Y-%m-%d %H:%M", time.localtime(_mtime(path))),
                    "note_count": note_count,
                })

    exports = []
    if os.path.isdir(EXPORT_DIR):
        for name in sorted(os.listdir(EXPORT_DIR), key=lambda n: _mtime(os.path.join(EXPORT_DIR, n)), reverse=True):
            path = os.path.join(EXPORT_DIR, name)
            if os.path.isdir(path):
                try:
                    md_count = len([f for f in os.listdir(path) if f.endswith(".md")])
                    jsonl_count = 1 if os.path.isfile(os.path.join(path, "notes.jsonl")) else 0
                except OSError:
                    md_count = jsonl_count = 0
                exports.append({
                    "name": name,
                    "path": path,
                    "mtime": _mtime(path),
                    "time": time.strftime("%Y-%m-%d %H:%M", time.localtime(_mtime(path))),
                    "md_count": md_count,
                    "jsonl": bool(jsonl_count),
                })

    # 达人导出（蒲公英）：蒲公英达人_*.xlsx + 同名 _完整数据.json 成组
    talent_exports = []
    if os.path.isdir(EXCEL_DIR):
        xlsx_names = [f for f in os.listdir(EXCEL_DIR)
                      if f.endswith(".xlsx") and f.startswith("蒲公英达人_") and not f.startswith(".")]
        for name in sorted(xlsx_names, key=lambda n: _mtime(os.path.join(EXCEL_DIR, n)), reverse=True):
            base = name[:-5]  # 去掉 .xlsx
            xlsx_path = os.path.join(EXCEL_DIR, name)
            json_path = os.path.join(EXCEL_DIR, f"{base}_完整数据.json")
            talent_exports.append({
                "name": name,
                "path": xlsx_path,
                "json_path": json_path if os.path.isfile(json_path) else None,
                "mtime": _mtime(xlsx_path),
                "time": time.strftime("%Y-%m-%d %H:%M", time.localtime(_mtime(xlsx_path))),
                "size": os.path.getsize(xlsx_path),
                "has_json": os.path.isfile(json_path),
            })

    # 热点分析（xhs-hotspot-analysis）：data/hotspot/<task_id>/
    hotspot_tasks = []
    if os.path.isdir(HOTSPOT_DIR):
        for name in sorted(os.listdir(HOTSPOT_DIR), key=lambda n: _mtime(os.path.join(HOTSPOT_DIR, n)), reverse=True):
            path = os.path.join(HOTSPOT_DIR, name)
            if not os.path.isdir(path) or name.startswith("."):
                continue
            sources = os.path.join(path, "sources.jsonl")
            analysis = os.path.join(path, "analysis.md")
            meta = os.path.join(path, "meta.json")
            query = ""
            note_count = 0
            if os.path.isfile(meta):
                try:
                    with open(meta, encoding="utf-8") as fh:
                        query = json.load(fh).get("query", "")
                except Exception:
                    pass
            if os.path.isfile(sources):
                try:
                    with open(sources, encoding="utf-8") as fh:
                        note_count = sum(1 for _ in fh)
                except Exception:
                    pass
            hotspot_tasks.append({
                "task_id": name,
                "path": path,
                "mtime": _mtime(path),
                "time": time.strftime("%Y-%m-%d %H:%M", time.localtime(_mtime(path))),
                "note_count": note_count,
                "has_analysis": os.path.isfile(analysis),
                "sources_path": sources,
                "analysis_path": analysis,
                "query": query,
            })

    return {
        "excel_dir": EXCEL_DIR,
        "media_dir": MEDIA_DIR,
        "export_dir": EXPORT_DIR,
        "excels": excels,
        "media_groups": media_groups,
        "exports": exports,
        "talent_exports": talent_exports,
        "hotspot_tasks": hotspot_tasks,
    }


def read_excel(path: str, limit: int = 200) -> dict:
    """读取 Excel 前几行内容，返回行列表（每行是 {header: value}）。"""
    if not os.path.isfile(path):
        raise ValueError("文件不存在")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    headers = None
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            headers = [str(c) if c is not None else "" for c in row]
            continue
        if idx > limit:
            break
        record = {}
        for h, v in zip(headers, row):
            if v is None:
                v = ""
            record[h] = str(v)
        rows.append(record)
    total = ws.max_row - 1 if ws.max_row else 0
    wb.close()
    return {
        "name": os.path.basename(path),
        "path": path,
        "headers": headers or [],
        "rows": rows,
        "total": total,
        "truncated": total > len(rows),
    }


def open_in_finder(path: str) -> dict:
    """在访达中打开文件所在目录并选中该文件；目录则直接打开。"""
    if not os.path.exists(path):
        raise ValueError("路径不存在")
    if sys.platform == "darwin":
        if os.path.isfile(path):
            subprocess.Popen(["open", "-R", path])
        else:
            subprocess.Popen(["open", path])
    elif sys.platform.startswith("win"):
        subprocess.Popen(["explorer", "/select,", path] if os.path.isfile(path) else ["explorer", path])
    else:
        subprocess.Popen(["xdg-open", os.path.dirname(path) if os.path.isfile(path) else path])
    return {"ok": True, "path": path}


def delete_path(path: str) -> dict:
    """删除 datas 或 data/hotspot 下的文件/文件夹（仅限这两个目录内，防止误删）。"""
    real = os.path.abspath(path)
    allowed = [os.path.abspath(DATAS_DIR), os.path.abspath(HOTSPOT_DIR)]
    ok_root = any(real.startswith(base + os.sep) for base in allowed)
    if not ok_root:
        raise ValueError("只能删除 datas/ 或 data/hotspot/ 目录内的文件")
    if real in allowed:
        raise ValueError("不能删除根目录")
    if not os.path.exists(real):
        raise ValueError("路径不存在")
    if os.path.isfile(real):
        os.remove(real)
    else:
        import shutil
        shutil.rmtree(real)
    return {"ok": True, "path": real}
